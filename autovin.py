#import necessary packages
import streamlit as st
import pandas as pd
import requests
import os
import openpyxl
import numpy as np
import json
from datetime import datetime
from io import BytesIO
import json

@st.cache_data

#create a fucntion that will group vehicles by type for the fleet summary, this will be called later within the 'confirm_vin()' function
def grouped_vehicles(dataframe):
    #iterate through the columns of the data frame
    for column in dataframe.columns:
        #remove none entries from the dataframe, replace None with empty string
        dataframe[column] = ['' if value is None else value for value in dataframe[column]]
    
    #iterate through the rows of the dataframe, create a list holding the NHTSA MAKE and NHTSA MODEL values returned from the NHTSA API, only include vehicles that are not lifts, trailers or unknown
    #we also do not record vehicles that do not have make or model recorded, if the VIN relates to a trailer or lift, record TRAILER or LIFT, if the vehicle type is unknown the vehicle type is
    #recorded as unconfirmed
    raw_vehicles = dataframe.apply(lambda row: row['NHTSA MAKE'] + ' ' + row['NHTSA MODEL'] if row['VEHICLE TYPE'] not in ['LIFT', 'TRAILER', 'UNKNOWN'] and row['NHTSA MAKE'] != '' and row['NHTSA MODEL'] != '' else 'TRAILER' if row['VEHICLE TYPE'] == 'TRAILER' else 'LIFT' if row['VEHICLE TYPE'] == 'LIFT' else 'UNCONFIRMED', axis=1).values.tolist()
    
    #create a list of the distinct makes and models
    distinct_vehicles = list(set(raw_vehicles))
    
    #count the number vehicles of a specific make and model
    counts = [raw_vehicles.count(item) for item in distinct_vehicles]
    
    #create a dictionary of the make/models and their vehicle count
    vehicles = {k: v for k, v in zip(distinct_vehicles, counts)}
    
    #record the vehicle, vin, make/model information of the invalid VINs as recorded by the account manager in the MCF deployment template
    add_unknown_info = dataframe.apply(lambda row: 'VEHICLE ' + '\n' + 'VIN: ' + row['VIN'] + '\n' + 'MAKE/MODEL: ' + row['MAKE'] + ' ' + row['MODEL'] if raw_vehicles[row.name] == 'UNCONFIRMED' else '', axis=1).tolist()
    
    #add split lines in between unknown vehicles for formatting purposes
    add_unknown_info = [value.splitlines() for value in add_unknown_info if value != '']
    
    #create a list to store the known vehicles
    known_vehicles = []
    
    #create a list to store unconfirmed vehicles, trailers and lifts, this will ensure trailer, lift, and unconfirmed vehicles are added to the end of the fleet summary
    bulk_vehicles = []
    
    #add vehicles to the vehicle lists
    for key, value in vehicles.items():
        #append the known_vehicles list with the dictinct vehicle makes/models and their vehicle count
        if key not in {'LIFT', 'TRAILER','UNCONFIRMED'}:
            known_vehicles.append(f'{key}: {value} \n')
        #append the bulk_vehicles list with lift, trailer and unconfirmed vehicle types and their respective vehicle counts
        else:
            bulk_vehicles.append(f'{key}: {value}\n')
    
    #sort the known vehicle list alphabetically and write it to known_output which is text that will later be output as the fleet summary
    known_output = ''.join(sorted(known_vehicles))
    #add the sorted bulk_vehicles list to the known_output text
    known_output += ''.join(sorted(bulk_vehicles))
    
    #create empty unknown output text
    unknown_output = ''
    
    #if there are any unconfirmed vehicles, add the related vehicle information to the unknown_output text
    if add_unknown_info != []:
        
        #format the output so the spaces between the VINs and Vehicles are even and the output is legible
        
        #create a variable to store length of the largest VIN
        length = 0
        
        #iterate through the unknown vehicle information
        for i, values in enumerate(add_unknown_info, start = 1):
            #add the relevant number to the Vehicle string, ex. first vehicle will say Vehicle 1 in the unknown vehicle output
            values[0] += str(i)
            #if any VIN is longer than 27 digits (which is a typo) cut the VIN to 27 digits and add '...' to indicate the VIN is cut off
            if len(values[1]) > 27:
                values[1].ljust(27)
                values[1] += '...'
            #if the length of the VIN is longer than the current length, update length
            if len(values[1]) > length:
                length = len(values[1])
        
        #add the unknown input to the unknown output, adjusting the length of all the VINs to the length of the longest VIN to ensure spacing is consistent and output is formatted nicely
        unknown_output = '\n'.join(f'{values[0]} INFO:    {values[1].ljust(length)}    {values[2]}' for values in add_unknown_info)
    
    #return the known vehicle output and the unknown vehicle output for the fleet summary
    return known_output, unknown_output
    

def confirm_vin(file_path):
    #some excel files have more than 1 sheet, we handle excel files with more than 1 sheet by telling the 
    #code to read the sheet named 'Vehicle & Asset List' as this is the standard naming convention
    #write the information from this sheet into dataframe named 'raw_vin_data'
    wb = openpyxl.load_workbook(file_path)
    res = len(wb.sheetnames)
    if res > 1:
        raw_vin_data = pd.read_excel(file_path, 'Vehicle & Asset List', header = 3) #turns excel into pandas dataframe
    else:
        raw_vin_data = pd.read_excel(file_path, header = 3)
        
    #assign new column names to raw_vin_data dataframe for dataframe to standardize raw_vin_data for query
    for column in raw_vin_data.columns:
        if 'vehicle asset name' in column.lower():
            raw_vin_data.rename(columns={column:'Vehicle Asset Name'}, inplace=True)
        elif 'model year' in column.lower():
            raw_vin_data.rename(columns={column:'Model Year'}, inplace=True)
        elif 'make' in column.lower():
            raw_vin_data.rename(columns={column:'Make'}, inplace=True)
        elif 'model' in column.lower():
            raw_vin_data.rename(columns={column:'Model'}, inplace=True)
        elif 'vin' in column.lower():
            raw_vin_data.rename(columns={column:'VIN'}, inplace=True)
        elif 'fuel type' in column.lower():
            raw_vin_data.rename(columns={column:'Fuel Type'}, inplace=True)
    
    #create base url that will be augmented with VIN for query
    base_url = 'https://vpic.nhtsa.dot.gov/api/vehicles/DecodeVin/'
    
    #create a dataframe for CAN compatability check in standard format with required columns
    vin_data = pd.DataFrame({'VRN':[], 'VIN': [], 'YEAR': [], 
                             'MAKE': [], 'MODEL': [], 'FUEL': [], 'COUNTRY': []})
    
    
    #write relevant info into vin_data dataframe using raw data from original sales document, only includes 
    #info from rows where vin has been entered, excludes NULL/NAN values
    for ind in raw_vin_data.index:
        if pd.isna(raw_vin_data['VIN'][ind]) == False:
            vin_data.loc[ind] = [raw_vin_data['Vehicle Asset Name'][ind],
                             raw_vin_data['VIN'][ind], raw_vin_data['Model Year'][ind],
                             raw_vin_data['Make'][ind], raw_vin_data['Model'][ind],
                             raw_vin_data['Fuel Type'][ind],
                             'US']
    
    #reset the vin dataframe index, index now begins at 0
    vin_data.reset_index(drop = True, inplace = True)
    
    #replace NAN/NULL values indicating empty cell with an empty string
    vin_data.replace(np.nan, '', inplace = True)
    
    #change the values in vin_data dataframe to strings, this is necessary for later string concatenation
    vin_data = vin_data.astype(str)
    
   #create list to store dictionaries, each dictionary will relate to a specific VIN or row of the dataframe
    results = []
    
    #extract VINs from vin_data dataframe into a list of values
    values = vin_data['VIN'].values.tolist()
    
    #create variable to keep track of which index is being used, this keeps track of what row of the dataframe
    #the code is on
    ind = 0
    
    #query the NHTSA VIN database using each VIN from the original sales document to collect info on vehicle 
    #year, make, model, fuel, and vehicle type, as MCF operates in United States all entries for Country = US
    
    #iterate through each VIN in list of VINs
    for value in values:
        #create variable indicating if a VIN has been corrected
        corrected = 'NO'
        #ensure the type of the VIN is string
        value = str(value)
        #handle common data entry errors 
        #remove spaces from VIN, accounts for common data entry error
        if ' ' in value:
            value = value.replace(" ", "")
            corrected = 'YES: Spaces Removed'
        #replace Q with 0
        if 'q' in value.lower():
            value = value.replace('Q','0')
            value = value.replace('q', '0')
            corrected = "YES: Replaced 'Q' with '0'"
        #replace O with 0
        if 'o' in value.lower() and 'unknown' not in value.lower():
            value = value.replace('O', '0')
            value = value.replace('o', '0')
            corrected = "YES: Replaced 'O' with '0'"
        #replace I with 1
        if 'i' in value.lower():
            value = value.replace('I', '1')
            value = value.replace('i', '1')
            corrected = "YES: Replaced 'I' with 1"
        #create VIN specific link to access details for API query
        url = base_url + value + '?format=json'
        #pulls details from url, bypasses certification verification error created by Michelin firewalls
        response = requests.get(url, verify = False)
        #check to see if vin is accurate, if accurate extract data into dictionary and add to results list
        try:
            #save url information as data variable for query
            data = response.json()
            #create key for decoding desired information from url data
            decoded_values = {item['Variable']: item['Value'] for item in data['Results']}
            #create a dictionary with vehicle information from VIN query, information based on specific VIN
            results.append({

                'VRN': vin_data['VRN'][ind],
                'VIN': value,
                'VIN CORRECTED': corrected,
                'NHTSA YEAR': decoded_values.get('Model Year', 'N/A'),
                'NHTSA MAKE': decoded_values.get('Make', 'N/A'),
                'NHTSA MODEL': decoded_values.get('Model', 'N/A'),
                'YEAR': vin_data['YEAR'][ind],
                'MAKE': vin_data['MAKE'][ind],
                'MODEL': vin_data['MODEL'][ind],
                'FUEL': decoded_values.get('Fuel Type - Primary', 'N/A'),
                'COUNTRY': 'US',
                'VEHICLE TYPE': decoded_values.get('Vehicle Type', 'N/A'),
                'ERROR CODE': decoded_values.get('Error Text', 'N/A')
            })
            #increase the index by 1, indicates code moves onto next VIN/row
            ind += 1
        #if vin not accurate, use error handling, will only move to this step if url produces empty response 
        #(data variable is empty)
        except json.JSONDecodeError as e:
            results.append({
                'VRN': vin_data['VRN'][ind],
                'VIN': value,
                'VIN CORRECTED': corrected,
                'NHTSA YEAR': 'Error',
                'NHTSA MAKE': 'Error',
                'NHTSA MODEL': 'Error',
                'YEAR': vin_data['YEAR'][ind],
                'MAKE': vin_data['MAKE'][ind],
                'MODEL': vin_data['MODEL'][ind],
                'FUEL': 'Error',
                'COUNTRY': 'US',
                'VEHICLE TYPE': 'Error',
                'ERROR CODE': 'Error: No information found for input VIN'
            })
            #increase the index by 1, indicates code moves onto next VIN/row
            ind += 1
        #if code times out, this error handling will make sure the code does not run indefinitely, if
        #encountered the code will stop processing VINs and communicate a time out error to the user
        except requests.exceptions.Timeout as e:
            return "Timed out"
            
    #create dataframe from list of dictionaries, each dictionary is a row within the 'results' dataframe
    results = pd.DataFrame(results)
    
    #create valid_vins dataframe that will be fed into CAN compatability check, exclude trailers, lifts
    #and invalid VINs, exclude all rows where primary fuel type is N/A, Error or None, such results indicate
    #valid vehicles require an energy source
    valid_vins = results[~results.FUEL.isin(['Not Applicable', 'Error', None])]
    
    #remove VEHICLE TYPE and ERROR CODE column from valid_vins dataframe to ensure in correct format for 
    #CAN compatability
    valid_vins.drop(['VIN CORRECTED', 'NHTSA YEAR','NHTSA MAKE', 'NHTSA MODEL', 'VEHICLE TYPE', 'ERROR CODE'], axis = 1, inplace = True)
    
    #remove duplicate VINs from CAN compatability check document
    valid_vins.drop_duplicates(subset=['VIN'],inplace= True)
    
    #create a list indicating if vehicles need to be checked manually, this will become a column in the
    #processed VINs doccument exported to the employee
    check_list = []
    
    #create a record of checked VINs, this is used later to check if there are duplicate VINs
    vins_checked = []
    
    #create a list of valid VINs appearing in the CAN compatable dataframe, these do not need to be manually
    #checked
    valid_vin_list = valid_vins['VIN'].values.tolist()
    
    ##add vehicle type information to vin_data dataframe for employee reference, added to help with manual checks
    ##vin_data = pd.concat([vin_data, results['VEHICLE TYPE']], axis = 1)
    
    ##determine if a manual check of a given vehicle vin is necessary
    ##iterate through results dataframe by index
    for ind in results.index:
        #check if the VIN is already in CAN dataframe and not a duplicate, if true no manual check
        if results['VIN'][ind] in valid_vin_list and results['VIN'][ind] not in vins_checked:
            check_list.append('NO')
        #if the vehicle is a trailer no manual checl
        elif results['VEHICLE TYPE'][ind] == 'TRAILER':
            check_list.append('NO')
        elif 'trailer' in results['MODEL'][ind].lower() or 'trailer' in results['VRN'][ind].lower():
            check_list.append('NO')
        #if vehicle is a lift, no manual check
        elif 'lift' in results['MODEL'][ind].lower() or 'lift' in results['VRN'][ind].lower():
            check_list.append('NO')
        elif 'example' in results['VIN'][ind].lower():
            check_list.append('NO')
        #if VIN is duplicate manual check is necessary
        elif results['VIN'][ind] in vins_checked:
            check_list.append('YES: Duplicate Vin')
        #otherwise a manual check is necessary
        else:
            check_list.append('YES')
        #add VIN to vins_checked list
        vins_checked.append(results['VIN'][ind])
        
    #update vehicle type to indicate the vehicle is a trailer, lift or type is unkown where necessary
    for ind in results.index:
        if results['VEHICLE TYPE'][ind] == None or results['VEHICLE TYPE'][ind] == 'Error':
                if 'trailer' in results['MODEL'][ind].lower():
                    results['VEHICLE TYPE'][ind] ='TRAILER'
                elif 'lift' in vin_data['MODEL'][ind].lower():
                    results['VEHICLE TYPE'][ind] = 'LIFT'
                else:
                #elif results['VEHICLE TYPE'][ind] == 'Error':
                    results['VEHICLE TYPE'][ind] = 'UNKNOWN'

    #create results column indicating that somone needs to manually check a vehicle's VIN info using check_list
    results.insert(len(results.columns) - 1, 'MANUAL CHECK NEEDED', check_list)
    
    #valid_vins should be written to a CSV that is uploaded to SalesForce CAN compatability check, file path
    #should be the same as the input file with _CAN appended
    CAN_file_path = os.path.splitext(file_path)[0] + "_CAN.csv"
    pd.DataFrame(valid_vins).to_csv(CAN_file_path, index = False)
    
    #create file path and name for the processed vins output file, this file path is the same as the input
    #file path with _processed appended
    processed_file_path = os.path.splitext(file_path)[0] + "_processed.xlsx"
    
    #write vin_data dataframe to Excel file table, this will be the inclusive excel file with all VINS,
    #error codes, manual checks and vehicle types for employee reference
    with pd.ExcelWriter(processed_file_path, engine='openpyxl') as writer:
        
        #create an Excel sheet names 'Processed VINs' to hold the dataframe
        results.to_excel(writer, index=False, sheet_name='Processed VINs')

        #access Excel file and worksheet 
        workbook = writer.book
        worksheet = writer.sheets['Processed VINs']

        #iterate through columns, find the max width of the cells in the column
        for idx, column in enumerate(worksheet.columns):
            #skip over 'ERROR CODE' column as it is the last column, formatting is unneccesary here
            if worksheet.cell(row=1, column=idx + 1).value != 'ERROR CODE':
                max_length = 0
                for cell in column:
                    if cell.value is not None:
                        max_length = max(max_length, len(str(cell.value)))
                    #adjust column width to show all data
                    adjusted_width = (max_length + 2)
                    worksheet.column_dimensions[chr(65 + idx)].width = adjusted_width  #use index to get column letter
            
            #adjust 'ERROR CODE' column to be the width of the title
            if worksheet.cell(row=1, column=idx + 1).value == 'ERROR CODE':
                worksheet.column_dimensions[chr(65 + idx)].width = 12
                
    known_vehicles, unknown_vehicles = grouped_vehicles(results)
    
    #save and return number of distinct vehicles, the processed excel and can csv file paths to export
    return known_vehicles, unknown_vehicles, processed_file_path, CAN_file_path

custom_css = """
    <style>
        @import url('https://fonts.googleapis.com/css2?family=Open+Sans&display=swap');

        body {
            font-family: 'Arial', 'Open Sans', sans-serif;
        }

        .custom-markdown {
            font-size: 16px;
            line-height: 1.5;
            max-width: 800px;
            width: 100%;
        }
        
        .custom-text-area {
            font-family: 'Arial', 'Open Sans', sans-serif;
            font-size: 16px;
            line-height: 1.5;
            padding: 10px;
            width: 100%;
            box-sizing: border-box;
            white-space: pre-wrap;
        }
        
        .larger-font {
            font-size: 18px;
            font-weight: bold;
        }
        
        .largest-font {
            font-size: 22px;
            font-weight: bold;
        }
        
        .title {
            font-family: 'Arial', 'Open Sans', sans-serif;
            font-size: 36px;
            font-weight: bold;
        }
        
    </style>
"""

st.markdown(custom_css, unsafe_allow_html=True)


#add the Michelin banner to the top of the application, if the image link breaks you can correct this by copying and
#pasting an alternative image url in the ()
st.image("https://www.tdtyres.com/wp-content/uploads/2018/12/kisspng-car-michelin-man-tire-logo-michelin-logo-5b4c286206fa03.5353854915317177300286.png")

#set the application title to 'Vin Decoder'
st.markdown('<div class="custom-text-area title">{}</div>'.format('VIN Decoder'), unsafe_allow_html=True)

#create a drag and drop box for file uploading, indicate that the file must be a CSV or Excel file
uploaded_file = st.file_uploader("Choose an Excel or CSV file", type=["xls", "xlsx", "csv"])

#check if session state vairables 'processed_file_path and can_file_path exist, checks if a file has been uploaded
#if variables do not exists assign None to variables
if "processed_file_path" not in st.session_state:
    st.session_state["processed_file_path"] = None
    st.session_state["can_file_path"] = None

#if a file hase been uplaoded begin processing the file
if uploaded_file is not None:
    with st.spinner('Processing...'):
        #label the input file path with the same name as the uploaded document
        input_file_path = uploaded_file.name
        #write the uploaded file to a disk
        with open(input_file_path, "wb") as f:
            f.write(uploaded_file.getbuffer())
        #call confirm vin to process the input file, save the returned file paths to export to the user
        known_vehicles, unknown_vehicles, processed_file_path, can_file_path = confirm_vin(input_file_path)
        #indicate to the user the processed excel file status
        st.session_state["processed_file_path"] = processed_file_path
        #indicate to the user the CAN csv file status
        st.session_state["can_file_path"] = can_file_path
        #tell the user that the file has been successfully processed
        st.success('File successfully processed!')

#check if CAN csv and processed excel file paths exist
if st.session_state["processed_file_path"] and st.session_state["can_file_path"]:
    with open(st.session_state["processed_file_path"], "rb") as f:
        processed_data = f.read()
    with open(st.session_state["can_file_path"], "rb") as f:
        can_data = f.read()
    #create button allowing user to download processed excel file
    st.download_button(
        label="Download Processed File",
        data=BytesIO(processed_data),
        file_name=os.path.basename(st.session_state["processed_file_path"]),
        mime='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    #create button allowing user to download CAN csv file
    st.download_button(
        label="Download CAN File",
        data=BytesIO(can_data),
        file_name=os.path.basename(st.session_state["can_file_path"]),
        mime='text/csv'
    )
    #st.markdown(custom_css, unsafe_allow_html=True)
    #box_height = (fleet_summary.count('\n') + 1) * 20
    
    
    
    st.markdown('<div class="custom-text-area largest-font">{}</div>'.format('Fleet Summary'), unsafe_allow_html=True)
    st.markdown('<div class="custom-text-area larger-font">{}</div>'.format('Vehicle Types'), unsafe_allow_html=True)
    st.markdown('<div class="custom-text-area">{}</div>'.format(known_vehicles), unsafe_allow_html=True)
    st.markdown('<div class="custom-text-area larger-font">{}</div>'.format('Unconfirmed Vehicles'), unsafe_allow_html=True)
    st.markdown('<div class="custom-text-area">{}</div>'.format(unknown_vehicles), unsafe_allow_html=True)

#document how to use the VIN decoder application to the user
st.markdown('<div class="custom-text-area largest-font">{}</div>'.format('User Guide'), unsafe_allow_html=True)

st.markdown('''This application checks customer VINs with the [National Highway Traffic Safety Administration API](https://vpic.nhtsa.dot.gov/api/) to confirm VIN accuracy. The API helps ensure the VINs are accurate and relate to relevant vehicles for the CAN compatibility check on Salesforce. This application can handle large volumes of VINs but greater numbers of uploaded VINs will slow down processing time. Processing 2200 VINs takes roughly 25 minutes. When uploading large numbers of VINs please be patient and do not close out the application while processing.''')

st.markdown('<div class="custom-text-area larger-font">{}</div>'.format('Input Document Requirements'), unsafe_allow_html=True)
            
    
st.markdown('''- The uploaded document containing the VINs must follow the standard [Michelin Connected Fleet Deployment Template.](https://michelingroup.sharepoint.com/:x:/s/DocumentLibrary/EeVf3pMJk4RMoqM5R17La4UBkXCvYKbbhiTalXbr-RIU9g?e=vxNr7V) This application cannot decipher different document formats. If an error is indicated with a file you upload, please check the uploaded document follows the formatting guidelines.
- Make sure the input document is not open on your computer. If the input document is open, a permission error will occur.
- The VIN column must include the VINs the user wants to query. This is the only field necessary to confirm the existence/accuracy of the VINs.
- The output documents will lack account information regarding the vehicle make, model, year, and fuel type if these input document columns are empty. 
- If you are interested in retrieving additional vehicle information from VINs alone please use the [Automated VIN Data Application](https://vindata.streamlit.app/).

***Example Input Document:*** [***VIN Example***](https://michelingroup.sharepoint.com/:x:/s/DocumentLibrary/EYifdfuMSAJAnSaoPxeselABySIDMB0nLNRKxhBfW1kHWQ?e=e3tOBv)

***Note:*** If you are interested in checking the accuracy/existence of VINs recorded in a different format/document: download the MCF Deployment Template linked above, then copy and paste the VINs into the VIN column and upload this document for bulk processing.''')

st.markdown('<div class="custom-text-area larger-font">{}</div>'.format('Fleet Summary Output'), unsafe_allow_html=True)
st.markdown('''- Provides information on vehicle type and number of vehicles.
- Used for customer fleet information confirmation.
- Unconfirmed vehicles are vehicles whose VINs do not exist within the NHTSA database indicating a VIN error and are not categorized as a trailer or lift.

***Note:*** If a VIN does relate to a trailer or lift but is listed as unconfirmed update the Model column of the VIN to include the word 'trailer' or 'lift' on the MCF Deployment Template. 
''')
            
st.markdown('<div class="custom-text-area larger-font">{}</div>'.format('Output File 1: CAN Compatibility Check'), unsafe_allow_html=True)

st.markdown('''- After comparison with the NHTSA VIN database, accurate and relevant VINs are written to a CSV file following the standard format for the CAN compatibility check. 
- VINs relating to trailers and lifts are considered irrelevant to the CAN compatibility check and are excluded from this document. 
- This CSV will have the same name as the original document followed by _CAN. This file includes VRN, Year, Make, Model, VIN, and Fuel Type information from the original input file.

***Example CAN Output Document:*** [***VIN Example_CAN***](https://michelingroup.sharepoint.com/:x:/s/DocumentLibrary/EacrWkHBryJNrWVnA9FilCQBwmIIHnSx5wraTDd4Whnm1g?e=RsNt07)
''')

st.markdown('<div class="custom-text-area larger-font">{}</div>'.format('Output File 2: Processed VINs'), unsafe_allow_html=True)

st.markdown('''- This secondary output file includes information on all VINs present in the original uploaded document, including VINs excluded from the CAN Compatibility Check document. 
- The application processes the original VIN document and determines the VIN's vehicle type, reports if a VIN was corrected, indicates whether a manual employee check for a VIN is necessary and provides error code information pertaining to the VIN.
- The 'VIN Corrected' column indicates if the VIN was corrected for common data entry errors. If the VIN did not need to be corrected for common data entry errors the 'VIN Corrected' column will say 'No.' If spaces were removed, O's and Q's were replaced with 0's or I's were replaced with 1's this will be indicated.
- An error code of 0 indicates there was no issue with the VIN. 
- A manual check is indicated as unnecessary if the VIN was considered valid and written to the CAN compatibility document or the vehicle type is a trailer or lift (irrelevant vehicle). 
- A manual check is necessary if the VIN was not written to the CAN compatibility file as a valid VIN and the VIN does not relate to a trailer or lift (could be a relevant vehicle). 
- This file will have the same name as the original document followed by _processed. This file also includes VRN, Year, Make, Model, VIN and Fuel Type information from the original document. 

***Example Processed Output Document:*** [***VIN Example_processed***](https://michelingroup.sharepoint.com/:x:/s/DocumentLibrary/EfORSzVsdVlMkvHwFupC0EgBnunZu8xgBLEsGDB0oX2kvA?e=pWE7N3)

If you are encountering issues with this application please contact the Service Excellence Team: MCFNAServiceExcellenceTeam@MichelinGroup.onmicrosoft.com
''')
